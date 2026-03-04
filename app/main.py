from io import BytesIO
from pathlib import Path

from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.responses import FileResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from fastapi import Request
from openpyxl import Workbook, load_workbook

from app.models import CheckResponse, Issue, RuleResult
from app.rules.registry import RULES, list_rule_meta

app = FastAPI(title="Excel Check Tool")

BASE_DIR = Path(__file__).resolve().parent
app.mount("/static", StaticFiles(directory=str(BASE_DIR / "static")), name="static")
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))


@app.get("/", response_class=HTMLResponse)
async def index(request: Request) -> HTMLResponse:
    return templates.TemplateResponse("index.html", {"request": request})


@app.get("/api/rules")
async def get_rules():
    return {"rules": [rule.model_dump() for rule in list_rule_meta()]}


def _count_levels(issues: list[Issue]) -> tuple[int, int, int]:
    error_count = sum(1 for i in issues if i.severity == "error")
    warning_count = sum(1 for i in issues if i.severity == "warning")
    info_count = sum(1 for i in issues if i.severity == "info")
    return error_count, warning_count, info_count


def _load_workbook_from_upload(file: UploadFile) -> Workbook:
    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 文件")

    data = file.file.read()
    if not data:
        raise HTTPException(status_code=400, detail="文件为空")

    try:
        workbook = load_workbook(filename=BytesIO(data), data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=400, detail=f"无法解析 Excel 文件: {exc}") from exc

    return workbook


@app.post("/api/check", response_model=CheckResponse)
async def check_excel(file: UploadFile = File(...), rule_ids: str = Form(...)):
    workbook = _load_workbook_from_upload(file)
    selected_ids = [item.strip() for item in rule_ids.split(",") if item.strip()]
    if not selected_ids:
        raise HTTPException(status_code=400, detail="请至少选择一个检查规则")

    invalid_rule_ids = [rid for rid in selected_ids if rid not in RULES]
    if invalid_rule_ids:
        raise HTTPException(status_code=400, detail=f"无效规则: {', '.join(invalid_rule_ids)}")

    summary: list[RuleResult] = []
    all_issues: list[Issue] = []

    for rid in selected_ids:
        rule = RULES[rid]
        issues = rule.checker(workbook)
        errors, warnings, infos = _count_levels(issues)
        summary.append(
            RuleResult(
                rule_id=rule.id,
                rule_name=rule.name,
                error_count=errors,
                warning_count=warnings,
                info_count=infos,
                issues=issues,
            )
        )
        all_issues.extend(issues)

    return CheckResponse(summary=summary, issues=all_issues)


@app.post("/api/report")
async def export_report(payload: CheckResponse):
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["Rule ID", "Rule Name", "Error", "Warning", "Info", "Total"])

    for item in payload.summary:
        total = item.error_count + item.warning_count + item.info_count
        summary_sheet.append(
            [
                item.rule_id,
                item.rule_name,
                item.error_count,
                item.warning_count,
                item.info_count,
                total,
            ]
        )

    issues_sheet = workbook.create_sheet("Issues")
    issues_sheet.append(["Rule ID", "Rule Name", "Severity", "Sheet", "Row", "Column", "Message"])
    for issue in payload.issues:
        issues_sheet.append(
            [
                issue.rule_id,
                issue.rule_name,
                issue.severity,
                issue.sheet,
                issue.row,
                issue.column,
                issue.message,
            ]
        )

    output_path = BASE_DIR / "report_output.xlsx"
    workbook.save(output_path)

    return FileResponse(
        str(output_path),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="excel_check_report.xlsx",
    )
